"""Spreadsheet reader — handles CSV (stdlib) and XLSX (openpyxl, lazy).

Install xlsx support with:  pip install ractogateway[rag-excel]
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any


def _require_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "Reading Excel (.xlsx/.xls) files requires the 'openpyxl' package. "
            "Install it with:  pip install ractogateway[rag-excel]"
        ) from exc
    return openpyxl


from ractogateway.rag._models.document import Document
from ractogateway.rag.readers.base import BaseReader

# Magic bytes: ZIP (PK header) — all XLSX/XLSM files are ZIP archives
_XLSX_MAGIC = b"PK\x03\x04"


class SpreadsheetReader(BaseReader):
    """Read CSV and Excel spreadsheets into plain text.

    Each row is rendered as a tab-separated line; an optional header row is
    prepended.  Multiple sheets in an XLSX workbook are separated by a
    ``--- Sheet: <name> ---`` divider.

    Accepts a file path (``str`` / ``Path``), raw ``bytes``, or any binary
    file-like object with a ``.read()`` method.  When bytes/buffer are
    provided, XLSX format is detected via the ZIP magic header
    (``PK\\x03\\x04``); everything else is treated as CSV/TSV.

    Parameters
    ----------
    max_rows:
        Maximum number of rows to read per sheet (``None`` = all).
    include_header:
        Whether to repeat the header row at the start of each sheet section.
    """

    def __init__(
        self,
        max_rows: int | None = None,
        include_header: bool = True,
    ) -> None:
        self._max_rows = max_rows
        self._include_header = include_header

    @property
    def supported_extensions(self) -> frozenset[str]:
        return frozenset({".csv", ".tsv", ".xlsx", ".xls"})

    # ------------------------------------------------------------------
    # Path-based entry point
    # ------------------------------------------------------------------

    def _read_path(self, path: Path) -> Document:
        ext = path.suffix.lower()
        if ext in {".csv", ".tsv"}:
            return self._read_csv_path(path)
        return self._read_xlsx_path(path)

    # ------------------------------------------------------------------
    # Bytes / buffer entry point
    # ------------------------------------------------------------------

    def _read_bytes(self, data: bytes, *, source_label: str = "<bytes>") -> Document:
        if data[:4] == _XLSX_MAGIC:
            return self._read_xlsx_bytes(data, source_label)
        return self._read_csv_bytes(data, source_label)

    # ------------------------------------------------------------------
    # CSV — path
    # ------------------------------------------------------------------

    def _read_csv_path(self, path: Path) -> Document:
        try:
            raw = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            raw = path.read_text(encoding="latin-1")

        rows = self._parse_csv_text(raw)
        return Document(
            content="\n".join("\t".join(row) for row in rows),
            source=str(path.resolve()),
            metadata={
                "extension": path.suffix.lower(),
                "filename": path.name,
                "size_bytes": path.stat().st_size,
                "row_count": len(rows),
                "col_count": max((len(r) for r in rows), default=0),
            },
        )

    # ------------------------------------------------------------------
    # CSV — bytes
    # ------------------------------------------------------------------

    def _read_csv_bytes(self, data: bytes, source_label: str) -> Document:
        try:
            raw = data.decode("utf-8")
        except UnicodeDecodeError:
            raw = data.decode("latin-1")

        rows = self._parse_csv_text(raw)
        return Document(
            content="\n".join("\t".join(row) for row in rows),
            source=source_label,
            metadata={
                "size_bytes": len(data),
                "row_count": len(rows),
                "col_count": max((len(r) for r in rows), default=0),
            },
        )

    def _parse_csv_text(self, raw: str) -> list[list[str]]:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=",\t;|")
        reader = csv.reader(io.StringIO(raw), dialect)
        rows = list(reader)
        if self._max_rows is not None:
            rows = rows[: self._max_rows + 1]  # +1 for header
        return rows

    # ------------------------------------------------------------------
    # XLSX — path
    # ------------------------------------------------------------------

    def _read_xlsx_path(self, path: Path) -> Document:
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames)
        sections, total_rows = self._parse_xlsx_workbook(wb)
        wb.close()
        return Document(
            content="\n\n".join(sections),
            source=str(path.resolve()),
            metadata={
                "extension": path.suffix.lower(),
                "filename": path.name,
                "size_bytes": path.stat().st_size,
                "sheet_count": len(sheet_names),
                "total_rows": total_rows,
            },
        )

    # ------------------------------------------------------------------
    # XLSX — bytes
    # ------------------------------------------------------------------

    def _read_xlsx_bytes(self, data: bytes, source_label: str) -> Document:
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
        sheet_names = list(wb.sheetnames)
        sections, total_rows = self._parse_xlsx_workbook(wb)
        wb.close()
        return Document(
            content="\n\n".join(sections),
            source=source_label,
            metadata={
                "size_bytes": len(data),
                "sheet_count": len(sheet_names),
                "total_rows": total_rows,
            },
        )

    def _parse_xlsx_workbook(self, wb: Any) -> tuple[list[str], int]:
        sections: list[str] = []
        total_rows = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if self._max_rows is not None and i >= self._max_rows:
                    break
                cells = [str(c) if c is not None else "" for c in row]
                rows_text.append("\t".join(cells))
                total_rows += 1
            if rows_text:
                sections.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows_text))
        return sections, total_rows
